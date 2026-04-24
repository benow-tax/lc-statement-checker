"""
내국신용장등전자발급명세서 검증 자동화 툴 — Streamlit 웹 앱
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  1차 검증: 전표 vs 홈택스 영세율세금계산서
  2차 검증: 전표 vs 홈택스 불러오기(구매확인서)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

import io
import streamlit as st
import pandas as pd
from datetime import datetime
from itertools import combinations

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────
# 페이지 설정
# ─────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="전자발급명세서 검증 자동화",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─────────────────────────────────────────────────────────────
# CSS 커스텀 스타일
# ─────────────────────────────────────────────────────────────

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;700;900&display=swap');

    html, body, [class*="css"] {
        font-family: 'Noto Sans KR', sans-serif;
    }

    /* 헤더 배너 */
    .main-header {
        background: linear-gradient(135deg, #1F3864 0%, #2E75B6 100%);
        color: white;
        padding: 2rem 2.5rem;
        border-radius: 16px;
        margin-bottom: 2rem;
        box-shadow: 0 8px 32px rgba(31,56,100,0.18);
    }
    .main-header h1 {
        font-size: 1.7rem;
        font-weight: 900;
        margin: 0 0 0.3rem 0;
        letter-spacing: -0.5px;
    }
    .main-header p {
        font-size: 0.9rem;
        opacity: 0.8;
        margin: 0;
    }

    /* 탭 카드 */
    .tab-card {
        background: #F8FAFC;
        border: 1.5px solid #E2E8F0;
        border-radius: 12px;
        padding: 1.5rem;
        margin-bottom: 1.2rem;
    }

    /* 배지 */
    .badge-1 {
        background: #1565C0;
        color: white;
        padding: 4px 14px;
        border-radius: 20px;
        font-size: 0.78rem;
        font-weight: 700;
        display: inline-block;
        margin-bottom: 0.8rem;
    }
    .badge-2 {
        background: #6A1B9A;
        color: white;
        padding: 4px 14px;
        border-radius: 20px;
        font-size: 0.78rem;
        font-weight: 700;
        display: inline-block;
        margin-bottom: 0.8rem;
    }

    /* 결과 요약 박스 */
    .result-box {
        border-radius: 10px;
        padding: 1rem 1.4rem;
        margin: 0.4rem 0;
        font-weight: 600;
        font-size: 0.95rem;
    }
    .result-ok   { background: #E8F5E9; color: #1B5E20; border-left: 4px solid #43A047; }
    .result-warn { background: #FFF8E1; color: #E65100; border-left: 4px solid #FFA726; }
    .result-diff { background: #FFEBEE; color: #B71C1C; border-left: 4px solid #E53935; }
    .result-grp  { background: #F3E5F5; color: #4A148C; border-left: 4px solid #8E24AA; }

    /* 구분선 */
    .divider {
        border: none;
        border-top: 1.5px solid #E2E8F0;
        margin: 1.5rem 0;
    }

    /* 안내 텍스트 */
    .guide-text {
        color: #64748B;
        font-size: 0.87rem;
        margin-bottom: 1rem;
    }

    /* Streamlit 기본 버튼 오버라이드 */
    div.stButton > button {
        width: 100%;
        padding: 0.65rem 1.5rem;
        font-size: 1rem;
        font-weight: 700;
        border-radius: 8px;
        border: none;
        transition: all 0.2s;
    }

    /* 다운로드 버튼 */
    div.stDownloadButton > button {
        width: 100%;
        padding: 0.65rem 1.5rem;
        font-size: 1rem;
        font-weight: 700;
        border-radius: 8px;
        background: #1F3864 !important;
        color: white !important;
        border: none;
        transition: all 0.2s;
    }

    /* 파일 업로더 라벨 */
    .stFileUploader label {
        font-weight: 600;
        color: #1E293B;
    }

    /* 탭 스타일 */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        background: #F1F5F9;
        padding: 6px;
        border-radius: 10px;
    }
    .stTabs [data-baseweb="tab"] {
        border-radius: 7px;
        font-weight: 600;
        font-size: 0.95rem;
        padding: 8px 24px;
    }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────
# 유틸리티
# ─────────────────────────────────────────────────────────────

GROUP_PALETTES = [
    {"bg": "E8F5E9", "font": "1B5E20", "border": "81C784"},
    {"bg": "E3F2FD", "font": "0D47A1", "border": "64B5F6"},
    {"bg": "F3E5F5", "font": "4A148C", "border": "CE93D8"},
    {"bg": "FFF8E1", "font": "E65100", "border": "FFD54F"},
    {"bg": "FCE4EC", "font": "880E4F", "border": "F48FB1"},
    {"bg": "E0F7FA", "font": "006064", "border": "4DD0E1"},
    {"bg": "FBE9E7", "font": "BF360C", "border": "FFAB91"},
    {"bg": "F9FBE7", "font": "33691E", "border": "C5E1A5"},
]
DEFAULT_L_BG = "FFF8E1"
DEFAULT_R_BG = "FFEBEE"


def norm_biz(x):
    return str(x).replace('-', '').replace(' ', '').strip().zfill(10)

def fmt_date(v):
    try:
        if pd.isnull(v): return ''
    except Exception: pass
    if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
    return str(v) if v else ''

def fmt_amt(v):
    try:
        if pd.isnull(v): return ''
        return int(v)
    except Exception: return v

def safe(v):
    try:
        if pd.isnull(v): return ''
    except Exception: pass
    return v

def thin_border(color="BDBDBD"):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border(color="888888"):
    thick = Side(style='medium', color=color)
    thin  = Side(style='thin',   color=color)
    return Border(left=thick, right=thick, top=thin, bottom=thin)

def make_header(ws, row, texts, bg, fg="FFFFFF", size=10):
    for col, t in enumerate(texts, 1):
        c = ws.cell(row=row, column=col, value=t)
        c.font = Font(bold=True, color=fg, name="Arial", size=size)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border()

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def is_summary_row(row, columns):
    for col in columns:
        val = row[col]
        try:
            if pd.isnull(val): continue
        except Exception: pass
        if str(val).strip() == '합계':
            return True
    return False

def split_summary(df):
    mask = df.apply(lambda row: is_summary_row(row, df.columns), axis=1)
    return df[~mask].copy().reset_index(drop=True), df[mask].copy().reset_index(drop=True)


# ─────────────────────────────────────────────────────────────
# 파일 로더
# ─────────────────────────────────────────────────────────────

def load_영세율(file_obj):
    df_probe = pd.read_excel(file_obj, nrows=3)
    cols = list(df_probe.columns)
    file_obj.seek(0)

    if '사업자등록번호' in cols and '금액(원)' in cols:
        df = pd.read_excel(file_obj)
        return df, '사업자등록번호', '금액(원)', '발급일', '거래처명', '서류번호', {7}
    else:
        df_raw = pd.read_excel(file_obj, header=None)
        header_row = None
        for i in range(min(10, len(df_raw))):
            vals = [str(v) for v in df_raw.iloc[i] if str(v) not in ['nan', 'NaT', 'None']]
            if '공급받는자사업자등록번호' in vals or '작성일자' in vals:
                header_row = i
                break
        if header_row is None:
            raise ValueError("영세율세금계산서 파일에서 헤더 행을 찾을 수 없습니다.")
        file_obj.seek(0)
        df = pd.read_excel(file_obj, header=header_row)
        df = df.rename(columns={
            '공급받는자사업자등록번호': '사업자등록번호',
            '합계금액': '금액(원)',
            '작성일자': '발급일',
            '상호.1': '거래처명',
            '승인번호': '서류번호',
        })
        if '거래처명' not in df.columns:
            for alt in ['상호.1', '상호1', '공급받는자상호']:
                if alt in df.columns:
                    df = df.rename(columns={alt: '거래처명'})
                    break
        df['금액(원)'] = pd.to_numeric(df['금액(원)'], errors='coerce')
        amt_col_idx = {list(df.columns).index('금액(원)') + 1}
        return df, '사업자등록번호', '금액(원)', '발급일', '거래처명', '서류번호', amt_col_idx


def load_구매확인서(file_obj):
    df = pd.read_excel(file_obj)
    cols = list(df.columns)

    if '사업자등록번호' in cols:
        biz_col = '사업자등록번호'
    elif '사업자번호' in cols:
        biz_col = '사업자번호'
    else:
        raise ValueError(f"구매확인서 파일에서 사업자번호 컬럼을 찾을 수 없습니다.\n컬럼 목록: {cols}")

    if '금액(원)' in cols:
        amt_col = '금액(원)'
        amt_cols_idx = {cols.index(amt_col) + 1}
    elif '원화환산금액' in cols:
        amt_col = '원화환산금액'
        amt_cols_idx = {cols.index(amt_col) + 1}
        if '원화금액' in cols:
            amt_cols_idx.add(cols.index('원화금액') + 1)
    else:
        raise ValueError(f"구매확인서 파일에서 금액 컬럼을 찾을 수 없습니다.\n컬럼 목록: {cols}")

    name_col = '거래처명' if '거래처명' in cols else ('상호' if '상호' in cols else (cols[4] if len(cols) > 4 else ''))
    date_col = '발급일' if '발급일' in cols else ('공급일' if '공급일' in cols else cols[3])
    doc_col  = '서류번호' if '서류번호' in cols else ('PKT번호' if 'PKT번호' in cols else cols[2])

    return df, biz_col, amt_col, date_col, name_col, doc_col, amt_cols_idx


# ─────────────────────────────────────────────────────────────
# 1:1 매칭
# ─────────────────────────────────────────────────────────────

def match_datasets(df_l, biz_l, amt_l, df_r, biz_r, amt_r):
    l = df_l.copy().reset_index(drop=True)
    r = df_r.copy().reset_index(drop=True)
    l['_biz'] = l[biz_l].apply(norm_biz)
    r['_biz'] = r[biz_r].apply(norm_biz)
    l['_amt'] = pd.to_numeric(l[amt_l], errors='coerce').fillna(0)
    r['_amt'] = pd.to_numeric(r[amt_r], errors='coerce').fillna(0)
    l['_key'] = l['_biz'] + '||' + l['_amt'].astype(str)
    r['_key'] = r['_biz'] + '||' + r['_amt'].astype(str)
    lg, rg = {}, {}
    for i, k in l['_key'].items(): lg.setdefault(k, []).append(i)
    for i, k in r['_key'].items(): rg.setdefault(k, []).append(i)
    ml, mr = set(), set()
    for key in set(lg) & set(rg):
        n = min(len(lg[key]), len(rg[key]))
        ml.update(lg[key][:n]); mr.update(rg[key][:n])
    return ml, mr, l[~l.index.isin(ml)].copy(), r[~r.index.isin(mr)].copy(), l, r, lg, rg


# ─────────────────────────────────────────────────────────────
# 합산 매칭
# ─────────────────────────────────────────────────────────────

def find_sum_groups(only_l, only_r, amt_l_col, amt_r_col, max_combo=5):
    ol  = only_l.copy().reset_index(drop=True)
    or_ = only_r.copy().reset_index(drop=True)
    if '_biz' not in ol.columns:  ol['_biz'] = ol.iloc[:, 3].apply(norm_biz)
    if '_biz' not in or_.columns: or_['_biz'] = or_.iloc[:, 3].apply(norm_biz)
    ol['_amt']  = pd.to_numeric(ol[amt_l_col],  errors='coerce').fillna(0)
    or_['_amt'] = pd.to_numeric(or_[amt_r_col], errors='coerce').fillna(0)

    used_l, used_r = set(), set()
    groups, palette_idx = [], 0

    # 전표 내 합산=0 (발행취소)
    for biz in ol['_biz'].unique():
        l_rows = ol[ol['_biz'] == biz]
        l_amts, l_idxs = l_rows['_amt'].tolist(), l_rows.index.tolist()
        if len(l_rows) < 2: continue
        for k in range(2, min(len(l_amts)+1, max_combo+1)):
            for combo_pos in combinations(range(len(l_amts)), k):
                li_combo = [l_idxs[p] for p in combo_pos]
                if any(li in used_l for li in li_combo): continue
                if sum(l_amts[p] for p in combo_pos) == 0:
                    pal = GROUP_PALETTES[palette_idx % len(GROUP_PALETTES)]
                    detail = ' + '.join(f"{int(l_amts[p]):,}원" for p in combo_pos)
                    groups.append({'palette': pal, 'l_idxs': li_combo, 'r_idxs': [],
                                   'note_l': f"🔄 발행취소 의심 (합산=0)\n{detail} = 0", 'note_r': ''})
                    used_l.update(li_combo); palette_idx += 1

    # 홈택스 내 합산=0 (발행취소)
    for biz in or_['_biz'].unique():
        r_rows = or_[or_['_biz'] == biz]
        r_amts, r_idxs = r_rows['_amt'].tolist(), r_rows.index.tolist()
        if len(r_rows) < 2: continue
        for k in range(2, min(len(r_amts)+1, max_combo+1)):
            for combo_pos in combinations(range(len(r_amts)), k):
                ri_combo = [r_idxs[p] for p in combo_pos]
                if any(ri in used_r for ri in ri_combo): continue
                if sum(r_amts[p] for p in combo_pos) == 0:
                    pal = GROUP_PALETTES[palette_idx % len(GROUP_PALETTES)]
                    detail = ' + '.join(f"{int(r_amts[p]):,}원" for p in combo_pos)
                    groups.append({'palette': pal, 'l_idxs': [], 'r_idxs': ri_combo,
                                   'note_l': '', 'note_r': f"🔄 발행취소 의심 (합산=0)\n{detail} = 0"})
                    used_r.update(ri_combo); palette_idx += 1

    # 전표 M건 ↔ 홈택스 N건 합산 일치
    for biz in set(ol['_biz'].unique()) & set(or_['_biz'].unique()):
        l_rows = ol[ol['_biz'] == biz]; r_rows = or_[or_['_biz'] == biz]
        l_amts, l_idxs = l_rows['_amt'].tolist(), l_rows.index.tolist()
        r_amts, r_idxs = r_rows['_amt'].tolist(), r_rows.index.tolist()

        for li, lamt in zip(l_idxs, l_amts):
            if li in used_l: continue
            found = False
            for k in range(2, min(len(r_amts)+1, max_combo+1)):
                if found: break
                for combo_pos in combinations(range(len(r_amts)), k):
                    ri_combo = [r_idxs[p] for p in combo_pos]
                    if any(ri in used_r for ri in ri_combo): continue
                    if sum(r_amts[p] for p in combo_pos) == lamt:
                        pal = GROUP_PALETTES[palette_idx % len(GROUP_PALETTES)]
                        r_detail = ' + '.join(f"{int(r_amts[p]):,}원" for p in combo_pos)
                        groups.append({'palette': pal, 'l_idxs': [li], 'r_idxs': ri_combo,
                                       'note_l': f"▶ 합산일치 (홈택스 {k}건)\n{r_detail}\n= {int(lamt):,}원",
                                       'note_r': f"▶ 합산일치 (전표 1건)\n전표 {int(lamt):,}원"})
                        used_l.add(li); used_r.update(ri_combo); palette_idx += 1; found = True; break

        for ri, ramt in zip(r_idxs, r_amts):
            if ri in used_r: continue
            found = False
            for k in range(2, min(len(l_amts)+1, max_combo+1)):
                if found: break
                for combo_pos in combinations(range(len(l_amts)), k):
                    li_combo = [l_idxs[p] for p in combo_pos]
                    if any(li in used_l for li in li_combo): continue
                    if sum(l_amts[p] for p in combo_pos) == ramt:
                        pal = GROUP_PALETTES[palette_idx % len(GROUP_PALETTES)]
                        l_detail = ' + '.join(f"{int(l_amts[p]):,}원" for p in combo_pos)
                        groups.append({'palette': pal, 'l_idxs': li_combo, 'r_idxs': [ri],
                                       'note_l': f"▶ 합산일치 (홈택스 1건)\n홈택스 {int(ramt):,}원",
                                       'note_r': f"▶ 합산일치 (전표 {k}건)\n{l_detail}\n= {int(ramt):,}원"})
                        used_r.add(ri); used_l.update(li_combo); palette_idx += 1; found = True; break

    style_l, style_r = {}, {}
    for g in groups:
        for li in g['l_idxs']: style_l[li] = {'palette': g['palette'], 'note': g['note_l']}
        for ri in g['r_idxs']: style_r[ri] = {'palette': g['palette'], 'note': g['note_r']}
    return groups, style_l, style_r


# ─────────────────────────────────────────────────────────────
# 합계 섹션 하단
# ─────────────────────────────────────────────────────────────

def write_summary_section(ws, start_row, summary_l, summary_r,
                           l_disp_cols, r_disp_cols, l_name, r_name, ncols,
                           auto_sum_r=None):
    nc_h      = len(l_disp_cols)
    amt_l_pos = {i+2 for i, c in enumerate(l_disp_cols) if '금액' in c or '원화' in c}
    amt_r_pos = {nc_h+1+i+2 for i, c in enumerate(r_disp_cols) if '금액' in c or '원화' in c}

    ws.row_dimensions[start_row].height = 6
    for col in range(1, ncols+1):
        c = ws.cell(row=start_row, column=col, value='')
        c.fill = PatternFill("solid", start_color="B0BEC5")
    start_row += 1

    ws.merge_cells(f"A{start_row}:{get_column_letter(ncols)}{start_row}")
    c = ws[f"A{start_row}"]
    c.value = "【 합계 】"
    c.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    c.fill = PatternFill("solid", start_color="263238")
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[start_row].height = 18
    start_row += 1

    def write_one(ws, rn, label, vals_dict, disp_cols, amt_pos, bg, offset):
        ws.row_dimensions[rn].height = 16
        c = ws.cell(row=rn, column=offset+1, value=label)
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        c.fill = PatternFill("solid", start_color="37474F")
        c.border = thin_border("B0BEC5")
        c.alignment = Alignment(vertical='center', horizontal='center')
        blank_fill = PatternFill("solid", start_color="F5F5F5")
        if offset == 0:
            for j in range(nc_h+2, ncols+1):
                c = ws.cell(row=rn, column=j, value=''); c.fill = blank_fill; c.border = thin_border("E0E0E0")
        else:
            for j in range(1, nc_h+2):
                c = ws.cell(row=rn, column=j, value=''); c.fill = blank_fill; c.border = thin_border("E0E0E0")
        for j, col in enumerate(disp_cols, offset+2):
            v = vals_dict.get(col, '')
            v2 = fmt_amt(v) if j in amt_pos else safe(v)
            c = ws.cell(row=rn, column=j, value=v2)
            c.font = Font(bold=True, name="Arial", size=9)
            c.fill = PatternFill("solid", start_color=bg)
            c.border = thin_border("B0BEC5")
            c.alignment = Alignment(vertical='center', horizontal='right' if j in amt_pos else 'center')
            if j in amt_pos and isinstance(v2, int): c.number_format = '#,##0'

    if not summary_l.empty:
        for _, srow in summary_l.iterrows():
            vals = {col: srow[col] if col in srow.index else '' for col in l_disp_cols}
            write_one(ws, start_row, f"{l_name} 합계", vals, l_disp_cols, amt_l_pos, "E8F0FE", 0)
            start_row += 1
    if not summary_r.empty:
        for _, srow in summary_r.iterrows():
            vals = {col: srow[col] if col in srow.index else '' for col in r_disp_cols}
            write_one(ws, start_row, f"{r_name} 합계", vals, r_disp_cols, amt_r_pos, "FCE8E6", nc_h+1)
            start_row += 1
    elif auto_sum_r:
        write_one(ws, start_row, f"{r_name} 합계 (자동계산)", auto_sum_r, r_disp_cols, amt_r_pos, "FCE8E6", nc_h+1)
        start_row += 1
    return start_row


# ─────────────────────────────────────────────────────────────
# 차이분석 행 렌더링
# ─────────────────────────────────────────────────────────────

def render_diff_row(ws, excel_row, data_row, disp_cols, amt_pos,
                    col_start, note_col, style_info, default_bg):
    if style_info:
        pal = style_info['palette']
        bg, fc, bdc, note = pal['bg'], pal['font'], pal['border'], style_info['note']
        bdr = thick_border(bdc)
    else:
        bg, fc, bdc, note = default_bg, "37474F", "E0E0E0", ''
        bdr = thin_border(bdc)

    fill = PatternFill("solid", start_color=bg)
    c = ws.cell(row=excel_row, column=col_start,
                value=(data_row.name + 1) if data_row is not None else '')
    c.fill = fill; c.font = Font(name="Arial", size=9, bold=bool(style_info))
    c.border = bdr; c.alignment = Alignment(vertical='center', horizontal='center')

    if data_row is not None:
        for j, col in enumerate(disp_cols, col_start+1):
            v = data_row[col]
            v2 = fmt_date(v) if ('일' in col and '금액' not in col and '원화' not in col) \
                 else fmt_amt(v) if j in amt_pos else safe(v)
            c = ws.cell(row=excel_row, column=j, value=v2)
            c.fill = fill
            c.font = Font(name="Arial", size=9, color=fc if style_info else "000000", bold=bool(style_info))
            c.border = bdr
            c.alignment = Alignment(vertical='center', horizontal='right' if j in amt_pos else 'left')
            if j in amt_pos and isinstance(v2, int): c.number_format = '#,##0'
        c = ws.cell(row=excel_row, column=note_col, value=note)
        c.fill = fill
        c.font = Font(name="Arial", size=8, color=fc if style_info else "000000", bold=bool(style_info))
        c.border = bdr
        c.alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)
    else:
        for j in range(col_start+1, note_col+1):
            c = ws.cell(row=excel_row, column=j, value='')
            c.fill = PatternFill("solid", start_color="FFFFFF"); c.border = thin_border("E0E0E0")


# ─────────────────────────────────────────────────────────────
# 엑셀 빌드
# ─────────────────────────────────────────────────────────────

def build_excel(file_title, df_l_raw, biz_l, amt_l, l_name,
                df_r_raw, biz_r, amt_r, r_name,
                l_disp_cols, l_disp_labels, r_disp_cols, r_disp_labels,
                raw_sheets, l_only_label, r_only_label):
    DARK = "1F3864"; MID = "2E75B6"
    S_FILL = {'✅ 일치': "E8F5E9", f'⚠️ {l_name}만': "FFF8E1", f'⚠️ {r_name}만': "FFEBEE"}
    S_FONT = {'✅ 일치': ("1B5E20", True), f'⚠️ {l_name}만': ("E65100", True), f'⚠️ {r_name}만': ("B71C1C", True)}

    df_l, summary_l = split_summary(df_l_raw)
    df_r, summary_r = split_summary(df_r_raw)

    if summary_r.empty:
        auto_sum_r = {}
        for col in r_disp_cols:
            if ('금액' in col or '원화' in col) and col in df_r.columns:
                auto_sum_r[col] = int(pd.to_numeric(df_r[col], errors='coerce').fillna(0).sum())
    else:
        auto_sum_r = None

    ml, mr, only_l, only_r, l_full, r_full, lg, rg = match_datasets(df_l, biz_l, amt_l, df_r, biz_r, amt_r)
    groups, style_l, style_r = find_sum_groups(only_l, only_r, amt_l, amt_r)

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    nc_h   = len(l_disp_labels)
    NOTE_L = nc_h + 2
    NOTE_R = NCOLS = (nc_h + 2) * 2
    R_START = nc_h + 3
    amt_l_pos = {i+2 for i, c in enumerate(l_disp_cols) if '금액' in c or '원화' in c}
    amt_r_pos = set()
    for j, col in enumerate(r_disp_cols, R_START+1):
        if '금액' in col or '원화' in col: amt_r_pos.add(j)

    def title_row(ws, text, ncols, bg=DARK):
        ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
        c = ws["A1"]; c.value = text
        c.font = Font(bold=True, size=13, color="FFFFFF", name="Arial")
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

    def stat_row(ws, n_ok, n_l, n_r, n_grp, ncols):
        ws.row_dimensions[2].height = 22
        sp = ncols // 4
        for rng, val, fg, bg in [
            (f"A2:{get_column_letter(sp)}2", f"✅ 일치: {n_ok}건", "2E7D32", "E8F5E9"),
            (f"{get_column_letter(sp+1)}2:{get_column_letter(sp*2)}2", f"⚠️ {l_name}에만: {n_l}건", "E65100", "FFF3CD"),
            (f"{get_column_letter(sp*2+1)}2:{get_column_letter(sp*3)}2", f"⚠️ {r_name}에만: {n_r}건", "B71C1C", "FFE0E0"),
            (f"{get_column_letter(sp*3+1)}2:{get_column_letter(ncols)}2", f"🔗 합산·취소일치: {n_grp}그룹", "4A148C", "F3E5F5"),
        ]:
            ws.merge_cells(rng); c = ws[rng.split(":")[0]]; c.value = val
            c.font = Font(bold=True, color=fg, size=10, name="Arial")
            c.fill = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal='center', vertical='center')

    # ── Sheet 1: 차이분석 ──────────────────────────────────────
    ws1 = wb.create_sheet("차이분석")
    title_row(ws1, file_title + " [차이 내역]", NCOLS)
    stat_row(ws1, len(ml), len(only_l), len(only_r), len(groups), NCOLS)
    ws1.row_dimensions[3].height = 18
    ws1.merge_cells(f"A3:{get_column_letter(NOTE_L)}3")
    ws1["A3"].value = f"【 {l_only_label} 】"
    ws1["A3"].font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    ws1["A3"].fill = PatternFill("solid", start_color="E65100")
    ws1["A3"].alignment = Alignment(horizontal='center', vertical='center')
    ws1.merge_cells(f"{get_column_letter(R_START)}3:{get_column_letter(NCOLS)}3")
    ws1[f"{get_column_letter(R_START)}3"].value = f"【 {r_only_label} 】"
    ws1[f"{get_column_letter(R_START)}3"].font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    ws1[f"{get_column_letter(R_START)}3"].fill = PatternFill("solid", start_color="B71C1C")
    ws1[f"{get_column_letter(R_START)}3"].alignment = Alignment(horizontal='center', vertical='center')
    make_header(ws1, 4, ["No"]+l_disp_labels+["비고"]+["No"]+r_disp_labels+["비고"], MID, size=9)
    ws1.row_dimensions[4].height = 18

    ld = only_l[l_disp_cols].reset_index(drop=True)
    rd = only_r[r_disp_cols].reset_index(drop=True)
    render_rows = []
    l_rendered, r_rendered = set(), set()
    for g in groups:
        max_len = max(len(g['l_idxs']), len(g['r_idxs']))
        for i in range(max_len):
            li = g['l_idxs'][i] if i < len(g['l_idxs']) else None
            ri = g['r_idxs'][i] if i < len(g['r_idxs']) else None
            render_rows.append((li, ri))
            if li is not None: l_rendered.add(li)
            if ri is not None: r_rendered.add(ri)
    remaining_l = [i for i in range(len(ld)) if i not in l_rendered]
    remaining_r = [i for i in range(len(rd)) if i not in r_rendered]
    for i in range(max(len(remaining_l), len(remaining_r))):
        render_rows.append((remaining_l[i] if i < len(remaining_l) else None,
                            remaining_r[i] if i < len(remaining_r) else None))

    for row_idx, (li, ri) in enumerate(render_rows):
        excel_row = row_idx + 5
        ws1.row_dimensions[excel_row].height = 36
        if li is not None:
            rd_row = ld.iloc[li].copy(); rd_row.name = li
            render_diff_row(ws1, excel_row, rd_row, l_disp_cols, amt_l_pos, 1, NOTE_L, style_l.get(li), DEFAULT_L_BG)
        else:
            for j in range(1, NOTE_L+1):
                c = ws1.cell(row=excel_row, column=j, value='')
                c.fill = PatternFill("solid", start_color="FFFFFF"); c.border = thin_border("E0E0E0")
        if ri is not None:
            rd_row = rd.iloc[ri].copy(); rd_row.name = ri
            render_diff_row(ws1, excel_row, rd_row, r_disp_cols, amt_r_pos, R_START, NOTE_R, style_r.get(ri), DEFAULT_R_BG)
        else:
            for j in range(R_START, NOTE_R+1):
                c = ws1.cell(row=excel_row, column=j, value='')
                c.fill = PatternFill("solid", start_color="FFFFFF"); c.border = thin_border("E0E0E0")

    ws1.freeze_panes = "A5"
    write_summary_section(ws1, len(render_rows)+5, summary_l, summary_r,
                          l_disp_cols, r_disp_cols, l_name, r_name, NCOLS, auto_sum_r=auto_sum_r)

    # ── Sheet 2: 전체매칭현황 ───────────────────────────────────
    ws2 = wb.create_sheet("전체매칭현황")
    title_row(ws2, file_title + " — 전체 통합 뷰", 12)
    make_header(ws2, 2,
        ["상태"]+[f"{l_name}_{lb}" for lb in l_disp_labels]+[""]+[f"{r_name}_{rb}" for rb in r_disp_labels],
        MID, size=9)
    ws2.row_dimensions[2].height = 18

    rows2 = []
    ml2, mr2 = set(), set()
    for key in set(lg) & set(rg):
        n2 = min(len(lg[key]), len(rg[key]))
        for li2, ri2 in zip(lg[key][:n2], rg[key][:n2]):
            rows2.append(('✅ 일치', [l_full.loc[li2,c] for c in l_disp_cols], [r_full.loc[ri2,c] for c in r_disp_cols]))
            ml2.add(li2); mr2.add(ri2)
    for li2 in l_full.index:
        if li2 not in ml2: rows2.append((f'⚠️ {l_name}만', [l_full.loc[li2,c] for c in l_disp_cols], ['']*len(r_disp_cols)))
    for ri2 in r_full.index:
        if ri2 not in mr2: rows2.append((f'⚠️ {r_name}만', ['']*len(l_disp_cols), [r_full.loc[ri2,c] for c in r_disp_cols]))
    rows2.sort(key=lambda x: 0 if '일치' in x[0] else 1)

    amt_l2 = {i+2 for i, c in enumerate(l_disp_cols) if '금액' in c or '원화' in c}
    amt_r2 = {len(l_disp_cols)+2+i+1 for i, c in enumerate(r_disp_cols) if '금액' in c or '원화' in c}
    sep = len(l_disp_cols) + 2
    for i, (status, lvals, rvals) in enumerate(rows2):
        rn = i+3; ws2.row_dimensions[rn].height = 16
        fill = PatternFill("solid", start_color=S_FILL.get(status, "FFFFFF"))
        fc2, fb2 = S_FONT.get(status, ("000000", False))
        c = ws2.cell(row=rn, column=1, value=status)
        c.font = Font(color=fc2, bold=fb2, name="Arial", size=9); c.fill = fill
        c.border = thin_border("E0E0E0"); c.alignment = Alignment(vertical='center', horizontal='center')
        for j, v in enumerate(lvals, 2):
            v2 = fmt_date(v) if isinstance(v, datetime) else fmt_amt(v) if j in amt_l2 else safe(v)
            c = ws2.cell(row=rn, column=j, value=v2); c.font = Font(name="Arial", size=9)
            c.fill = fill; c.border = thin_border("E0E0E0")
            c.alignment = Alignment(vertical='center', horizontal='right' if j in amt_l2 else 'left')
            if j in amt_l2 and isinstance(v2, int): c.number_format = '#,##0'
        ws2.cell(row=rn, column=sep, value='').fill = fill
        ws2.cell(row=rn, column=sep).border = thin_border("E0E0E0")
        for j, v in enumerate(rvals, sep+1):
            v2 = fmt_date(v) if isinstance(v, datetime) else fmt_amt(v) if j in amt_r2 else safe(v)
            c = ws2.cell(row=rn, column=j, value=v2); c.font = Font(name="Arial", size=9)
            c.fill = fill; c.border = thin_border("E0E0E0")
            c.alignment = Alignment(vertical='center', horizontal='right' if j in amt_r2 else 'left')
            if j in amt_r2 and isinstance(v2, int): c.number_format = '#,##0'
    ws2.freeze_panes = "A3"
    write_summary_section(ws2, len(rows2)+3, summary_l, summary_r,
                          l_disp_cols, r_disp_cols, l_name, r_name, 12, auto_sum_r=auto_sum_r)

    # ── Raw 시트 ────────────────────────────────────────────────
    raw_cfg = [("37474F","546E7A","ECEFF1","CFD8DC"), ("1A237E","283593","E8EAF6","C5CAE9"), ("4A148C","6A1B9A","F3E5F5","CE93D8")]
    for idx, (sheet_name, df_raw, amt_cols) in enumerate(raw_sheets):
        tc, hc, ac, bc = raw_cfg[idx % 3]
        raw_sheet_name = sheet_name.replace("원본_", "raw_")
        ws_r = wb.create_sheet(raw_sheet_name); nc = len(df_raw.columns)
        ws_r.merge_cells(f"A1:{get_column_letter(nc)}1")
        ws_r["A1"].value = raw_sheet_name.replace("raw_", "raw 데이터 — ")
        ws_r["A1"].font = Font(bold=True, size=12, color="FFFFFF", name="Arial")
        ws_r["A1"].fill = PatternFill("solid", start_color=tc)
        ws_r["A1"].alignment = Alignment(horizontal='center', vertical='center')
        ws_r.row_dimensions[1].height = 24
        make_header(ws_r, 2, list(df_raw.columns), hc, size=9)
        has_summary = any(is_summary_row(row, df_raw.columns) for _, row in df_raw.iterrows())
        for i, row in df_raw.iterrows():
            rn = i+3; row_is_summ = is_summary_row(row, df_raw.columns)
            bg = "263238" if row_is_summ else ("FFFFFF" if i % 2 == 0 else ac)
            for j, v in enumerate(row, 1):
                v2 = fmt_date(v) if isinstance(v, datetime) else fmt_amt(v) if j in amt_cols else safe(v)
                c = ws_r.cell(row=rn, column=j, value=v2)
                c.fill = PatternFill("solid", start_color=bg)
                c.font = Font(name="Arial", size=9, bold=row_is_summ, color="FFFFFF" if row_is_summ else "000000")
                c.border = thin_border(bc)
                c.alignment = Alignment(vertical='center', horizontal='right' if j in amt_cols else 'left')
                if j in amt_cols and isinstance(v2, int): c.number_format = '#,##0'
        if not has_summary:
            summ_row = len(df_raw)+3; ws_r.row_dimensions[summ_row].height = 18
            for j in range(1, nc+1):
                c = ws_r.cell(row=summ_row, column=j)
                if j == 1: c.value = "합계 (자동계산)"
                elif j in amt_cols:
                    total = pd.to_numeric(df_raw.iloc[:, j-1], errors='coerce').fillna(0).sum()
                    c.value = int(total); c.number_format = '#,##0'
                else: c.value = ''
                c.fill = PatternFill("solid", start_color="263238")
                c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
                c.border = thin_border(bc)
                c.alignment = Alignment(vertical='center',
                                        horizontal='right' if j in amt_cols else ('center' if j == 1 else 'left'))
        ws_r.freeze_panes = "A3"

    set_col_widths(wb.worksheets[0], [5,12,24,14,13,14,30, 5,12,24,14,13,22,30])
    set_col_widths(wb.worksheets[1], [13,12,26,14,13,14,3,12,26,14,13,22])

    return wb, len(ml), len(only_l), len(only_r), len(groups)


# ─────────────────────────────────────────────────────────────
# Streamlit UI
# ─────────────────────────────────────────────────────────────

st.markdown("""
<div class="main-header">
    <h1>📋 내국신용장등전자발급명세서 검증 자동화</h1>
    <p>전표 ↔ 홈택스 데이터 자동 비교 · 차이 내역 추출 · 합산 일치 그룹 자동 탐지</p>
</div>
""", unsafe_allow_html=True)

# ── 사용 가이드 (접을 수 있는 섹션) ──────────────────────────
with st.expander("📖 데이터 준비 가이드 — 파일을 어디서 뽑나요?", expanded=False):
    st.markdown("""
<style>
.guide-section {
    display: flex;
    gap: 1.2rem;
    margin-bottom: 1rem;
    flex-wrap: wrap;
}
.guide-card {
    flex: 1;
    min-width: 220px;
    border-radius: 12px;
    padding: 1.2rem 1.4rem;
    border: 1.5px solid #E2E8F0;
    background: #FAFAFA;
}
.guide-card h4 {
    margin: 0 0 0.6rem 0;
    font-size: 0.95rem;
    font-weight: 700;
}
.guide-card .step {
    display: flex;
    align-items: flex-start;
    gap: 0.5rem;
    margin-bottom: 0.45rem;
    font-size: 0.84rem;
    color: #374151;
    line-height: 1.5;
}
.guide-card .num {
    background: #1F3864;
    color: white;
    border-radius: 50%;
    width: 18px;
    height: 18px;
    min-width: 18px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 0.7rem;
    font-weight: 700;
    margin-top: 2px;
}
.guide-card.blue  { border-color: #93C5FD; background: #EFF6FF; }
.guide-card.blue h4  { color: #1D4ED8; }
.guide-card.blue .num { background: #1D4ED8; }
.guide-card.purple { border-color: #C4B5FD; background: #F5F3FF; }
.guide-card.purple h4 { color: #6D28D9; }
.guide-card.purple .num { background: #6D28D9; }
.guide-card.green  { border-color: #6EE7B7; background: #ECFDF5; }
.guide-card.green h4  { color: #065F46; }
.guide-card.green .num { background: #065F46; }
.guide-note {
    background: #FFF8E1;
    border-left: 4px solid #FFA726;
    border-radius: 0 8px 8px 0;
    padding: 0.7rem 1rem;
    font-size: 0.84rem;
    color: #78350F;
    margin-top: 0.5rem;
}
</style>

<div class="guide-section">

  <div class="guide-card blue">
    <h4>📁 전표 파일 (1차·2차 공통)</h4>
    <div class="step"><div class="num">1</div><span>ERP 접속 후 <b>내국신용장등전자발급명세서</b> 메뉴 이동</span></div>
    <div class="step"><div class="num">2</div><span>조회 기간 설정 후 데이터 조회</span></div>
    <div class="step"><div class="num">3</div><span>조회 결과를 <b>엑셀로 내보내기</b> (합계 행 포함)</span></div>
  </div>

  <div class="guide-card purple">
    <h4>📁 홈택스 영세율세금계산서 파일 (1차용)</h4>
    <div class="step"><div class="num">1</div><span><b>홈택스</b> 접속 → 로그인</span></div>
    <div class="step"><div class="num">2</div><span><b>조회/발급 → 세금계산서 → 매출 세금계산서</b> 메뉴 이동</span></div>
    <div class="step"><div class="num">3</div><span>기간 설정 후 <b>영세율</b> 세금계산서만 필터 조회</span></div>
    <div class="step"><div class="num">4</div><span>조회 결과 <b>엑셀 다운로드</b></span></div>
  </div>

  <div class="guide-card green">
    <h4>📁 홈택스 불러오기(구매확인서) 파일 (2차용)</h4>
    <div class="step"><div class="num">1</div><span>ERP 접속 후 <b>내국신용장등전자발급명세서</b> 메뉴 이동</span></div>
    <div class="step"><div class="num">2</div><span><b>홈택스 불러오기</b> 버튼 클릭</span></div>
    <div class="step"><div class="num">3</div><span>불러온 구매확인서 데이터를 <b>엑셀로 내보내기</b></span></div>
  </div>

</div>

<div class="guide-note">
  💡 <b>검증 순서 안내</b> &nbsp;|&nbsp;
  <b>1차:</b> 전표 파일 + 홈택스 영세율세금계산서 파일로 비교 → 차이 확인 후 전표 수정
  &nbsp;&nbsp;→&nbsp;&nbsp;
  <b>2차:</b> 수정된 전표 파일 + 홈택스 불러오기(구매확인서) 파일로 최종 비교
</div>
""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

tab1, tab2 = st.tabs(["  1차 검증  |  전표 vs 홈택스 영세율세금계산서  ", "  2차 검증  |  전표 vs 홈택스 불러오기(구매확인서)  "])
if tab1 is None or tab2 is None:
    st.stop()

# ══════════════════════════════════════════════════════════════
# 1차 검증 탭
# ══════════════════════════════════════════════════════════════
with tab1:
    st.markdown('<span class="badge-1">1차 검증</span>', unsafe_allow_html=True)
    st.markdown('<p class="guide-text">전표와 홈택스 영세율세금계산서 파일을 업로드하면 자동으로 비교 분석합니다.</p>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        f1_전표 = st.file_uploader("📄 전표 파일", type=["xlsx", "xls"], key="f1_전표",
                                    help="전표 엑셀 파일을 업로드하세요")
    with col2:
        f1_영세율 = st.file_uploader("📄 홈택스 영세율세금계산서 파일", type=["xlsx", "xls"], key="f1_영세율",
                                      help="홈택스에서 조회한 영세율세금계산서 파일을 업로드하세요")

    st.markdown('<hr class="divider">', unsafe_allow_html=True)

    btn1 = st.button("▶  1차 검증 실행", key="btn1", type="primary", use_container_width=True)

    if btn1:
        if not f1_전표 or not f1_영세율:
            st.error("⚠️ 전표 파일과 홈택스 영세율세금계산서 파일을 모두 업로드해 주세요.")
        else:
            with st.spinner("데이터 매칭 및 합산 분석 중..."):
                try:
                    df_전표 = pd.read_excel(f1_전표)
                    df_영세율, biz_r, amt_r, date_r, name_r, doc_r, amt_cols_r = load_영세율(f1_영세율)

                    r_avail = [c for c in [date_r, name_r, biz_r, amt_r, doc_r] if c in df_영세율.columns]
                    r_labels_map = {date_r:"발급일", name_r:"거래처명", biz_r:"사업자등록번호", amt_r:"금액(원)", doc_r:"서류번호"}

                    wb, n_ok, n_l, n_r, n_grp = build_excel(
                        "1차 검증: 전표 vs 홈택스 영세율세금계산서",
                        df_전표, '사업자등록번호', '금액(원)', '전표',
                        df_영세율, biz_r, amt_r, '홈택스영세율',
                        ['발급일','거래처명','사업자등록번호','금액(원)','비고'],
                        ["발급일","거래처명","사업자등록번호","금액(원)","비고"],
                        r_avail, [r_labels_map.get(c,c) for c in r_avail],
                        [("원본_전표", df_전표, {7}), ("원본_홈택스_영세율세금계산서", df_영세율, amt_cols_r)],
                        "전표에만 있는 내역 (홈택스 미등록 의심)",
                        "홈택스 영세율에만 있는 내역 (전표 미입력 의심)",
                    )

                    # 결과 요약
                    st.markdown("### 검증 결과")
                    c1, c2, c3, c4 = st.columns(4)
                    with c1: st.markdown(f'<div class="result-box result-ok">✅ 일치<br><b style="font-size:1.4rem">{n_ok}건</b></div>', unsafe_allow_html=True)
                    with c2: st.markdown(f'<div class="result-box result-warn">⚠️ 전표에만<br><b style="font-size:1.4rem">{n_l}건</b></div>', unsafe_allow_html=True)
                    with c3: st.markdown(f'<div class="result-box result-diff">⚠️ 홈택스에만<br><b style="font-size:1.4rem">{n_r}건</b></div>', unsafe_allow_html=True)
                    with c4: st.markdown(f'<div class="result-box result-grp">🔗 합산·취소일치<br><b style="font-size:1.4rem">{n_grp}그룹</b></div>', unsafe_allow_html=True)

                    # 엑셀 다운로드
                    st.markdown("<br>", unsafe_allow_html=True)
                    buf = io.BytesIO()
                    wb.save(buf); buf.seek(0)
                    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                    st.download_button(
                        label="⬇️  결과 엑셀 다운로드",
                        data=buf.getvalue(),
                        file_name=f"[1차검증]전표vs홈택스영세율_{ts}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

                except Exception as e:
                    st.error(f"오류가 발생했습니다:\n\n{e}")

# ══════════════════════════════════════════════════════════════
# 2차 검증 탭
# ══════════════════════════════════════════════════════════════
with tab2:
    st.markdown('<span class="badge-2">2차 검증</span>', unsafe_allow_html=True)
    st.markdown('<p class="guide-text">1차 수정 완료된 전표와 홈택스 불러오기(구매확인서) 파일을 업로드하면 자동으로 비교 분석합니다.</p>', unsafe_allow_html=True)

    col3, col4 = st.columns(2)
    with col3:
        f2_전표 = st.file_uploader("📄 전표 파일 (수정 후)", type=["xlsx", "xls"], key="f2_전표",
                                    help="1차 수정이 완료된 전표 파일을 업로드하세요")
    with col4:
        f2_구매 = st.file_uploader("📄 홈택스 불러오기(구매확인서) 파일", type=["xlsx", "xls"], key="f2_구매",
                                    help="홈택스에서 불러온 구매확인서 파일을 업로드하세요")

    st.markdown('<hr class="divider">', unsafe_allow_html=True)

    btn2 = st.button("▶  2차 검증 실행", key="btn2", type="primary", use_container_width=True)

    if btn2:
        if not f2_전표 or not f2_구매:
            st.error("⚠️ 전표 파일과 홈택스 불러오기(구매확인서) 파일을 모두 업로드해 주세요.")
        else:
            with st.spinner("데이터 매칭 및 합산 분석 중..."):
                try:
                    df_전표2 = pd.read_excel(f2_전표)
                    df_구매, biz_r2, amt_r2, date_r2, name_r2, doc_r2, amt_cols_r2 = load_구매확인서(f2_구매)

                    r_avail2 = [c for c in [date_r2, name_r2, biz_r2, amt_r2, doc_r2] if c in df_구매.columns]
                    r_labels2 = {date_r2:"발급일", name_r2:"거래처명", biz_r2:"사업자등록번호", amt_r2:"금액(원)", doc_r2:"서류번호"}

                    wb2, n_ok2, n_l2, n_r2, n_grp2 = build_excel(
                        "2차 검증: 전표 vs 홈택스 불러오기(구매확인서)",
                        df_전표2, '사업자등록번호', '금액(원)', '전표',
                        df_구매, biz_r2, amt_r2, '홈택스불러오기(구매확인서)',
                        ['발급일','거래처명','사업자등록번호','금액(원)','비고'],
                        ["발급일","거래처명","사업자등록번호","금액(원)","비고"],
                        r_avail2, [r_labels2.get(c,c) for c in r_avail2],
                        [("원본_전표", df_전표2, {7}), ("원본_홈택스_불러오기(구매확인서)", df_구매, amt_cols_r2)],
                        "전표에만 있는 내역 (홈택스 미등록 의심)",
                        "홈택스 불러오기(구매확인서)에만 있는 내역 (전표 미입력 의심)",
                    )

                    st.markdown("### 검증 결과")
                    c1, c2, c3, c4 = st.columns(4)
                    with c1: st.markdown(f'<div class="result-box result-ok">✅ 일치<br><b style="font-size:1.4rem">{n_ok2}건</b></div>', unsafe_allow_html=True)
                    with c2: st.markdown(f'<div class="result-box result-warn">⚠️ 전표에만<br><b style="font-size:1.4rem">{n_l2}건</b></div>', unsafe_allow_html=True)
                    with c3: st.markdown(f'<div class="result-box result-diff">⚠️ 홈택스에만<br><b style="font-size:1.4rem">{n_r2}건</b></div>', unsafe_allow_html=True)
                    with c4: st.markdown(f'<div class="result-box result-grp">🔗 합산·취소일치<br><b style="font-size:1.4rem">{n_grp2}그룹</b></div>', unsafe_allow_html=True)

                    st.markdown("<br>", unsafe_allow_html=True)
                    buf2 = io.BytesIO()
                    wb2.save(buf2); buf2.seek(0)
                    ts2 = datetime.now().strftime("%Y%m%d_%H%M%S")
                    st.download_button(
                        label="⬇️  결과 엑셀 다운로드",
                        data=buf2.getvalue(),
                        file_name=f"[2차검증]전표vs홈택스불러오기(구매확인서)_{ts2}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

                except Exception as e:
                    st.error(f"오류가 발생했습니다:\n\n{e}")
